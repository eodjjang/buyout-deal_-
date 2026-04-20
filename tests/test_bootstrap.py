from __future__ import annotations

from openpyxl import Workbook

from lbo_template.layout import ALL_SHEETS


def test_build_returns_workbook(wb: Workbook) -> None:
    assert isinstance(wb, Workbook)


def test_default_sheet_removed(wb: Workbook) -> None:
    assert "Sheet" not in wb.sheetnames


def test_all_13_sheets_created_in_correct_order(wb: Workbook) -> None:
    assert wb.sheetnames == ALL_SHEETS
    assert len(wb.sheetnames) == 13
