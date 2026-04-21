from openpyxl.styles import Font, PatternFill
from lbo_template import conventions as c


def test_color_constants():
    assert c.COLOR_SECTION_HEADER_FILL == "1F4E79"
    assert c.COLOR_COLUMN_HEADER_FILL == "D9E1F2"
    assert c.COLOR_INPUT_FILL == "F2F2F2"
    assert c.COLOR_KEY_OUTPUT_FILL == "BDD7EE"
    assert c.COLOR_INPUT_FONT == "0000FF"
    assert c.COLOR_CALC_FONT == "000000"
    assert c.COLOR_SAMETAB_LINK_FONT == "800080"
    assert c.COLOR_CROSSTAB_LINK_FONT == "008000"
    assert c.COLOR_CIQ_FORMULA_FONT == "008B8B"


def test_number_formats():
    assert c.NUM_FMT_ACCOUNTING == '#,##0;(#,##0);"-"'
    assert c.NUM_FMT_PERCENT == "0.0%"
    assert c.NUM_FMT_MULTIPLE == '0.0"x"'
    assert c.NUM_FMT_BPS == '0" bp"'


def test_style_factories_return_correct_types():
    assert isinstance(c.section_header_font(), Font)
    assert isinstance(c.input_font(), Font)
    assert isinstance(c.section_header_fill(), PatternFill)
    assert isinstance(c.input_fill(), PatternFill)


def test_fy_axis_columns():
    assert c.FY_AXIS_COLUMNS == ["B", "C", "D", "E", "F", "G", "H", "I"]
    assert c.FY_AXIS_LABELS == ["FY-2", "FY-1", "LTM", "FY1", "FY2", "FY3", "FY4", "FY5"]


def test_define_name_registers_workbook_scoped_name():
    from openpyxl import Workbook

    wb = Workbook()
    c.define_name(wb, "My_Anchor", "'Sheet'!$B$3")
    assert "My_Anchor" in wb.defined_names
    assert wb.defined_names["My_Anchor"].attr_text == "'Sheet'!$B$3"
