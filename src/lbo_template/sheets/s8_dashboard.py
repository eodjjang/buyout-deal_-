"""8_Dashboard — Word-paste-ready summary tables + DASH_* named range cluster."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment

from lbo_template.layout import SHEET_DASH, SHEET_DEBT, SHEET_INPUT, SHEET_OVERLAY, SHEET_RETURNS
from lbo_template import conventions as c
from lbo_template.sheets.s3_overlay import STRESSED_EBITDA_ROW, STRESSED_CAPEX_ROW

# Overlay FY columns (FY_AXIS: E..I = FY1..FY5).
OVERLAY_FY_COLS = ["E", "F", "G", "H", "I"]

# 표 4 — 차주 자금수지표 (상세): 헤더 row 27, 데이터 28..53
T4_HEADER_ROW = 27
T4_ROW_기초 = 28
T4_ROW_인수관련 = 29
T4_ROW_인수금융조달 = 30
T4_ROW_재무대책 = 31
T4_ROW_인수대금지급 = 32
T4_ROW_영업활동 = 33
T4_ROW_EBITDA = 34
T4_ROW_투자활동 = 35
T4_ROW_CAPEX = 36
T4_ROW_배당수익 = 37
T4_ROW_원리금상환재원 = 38
T4_ROW_추가차입 = 39
T4_ROW_기존차입원리금 = 40
T4_ROW_기존원금 = 41
T4_ROW_기존이자 = 42
T4_ROW_배당금지급 = 43
T4_ROW_법인세기타 = 44
T4_ROW_본건상환재원 = 45
T4_ROW_본건이자 = 46
T4_ROW_TrA = 47
T4_ROW_TrB = 48
T4_ROW_TrC = 49
T4_ROW_Tr원금합계 = 50  # DASH_CFTable Row7 매핑 (= Tr.A+B+C)
T4_ROW_원리금이후 = 51
T4_ROW_리파이낸싱 = 52
T4_ROW_기말 = 53

DASH_CF_SUMMARY_ROWS = {
    1: T4_ROW_기초,
    2: T4_ROW_EBITDA,
    3: T4_ROW_CAPEX,
    4: T4_ROW_배당수익,
    5: T4_ROW_기존차입원리금,
    6: T4_ROW_본건이자,
    7: T4_ROW_Tr원금합계,
    8: T4_ROW_기말,
}


def _table_title(ws: Worksheet, cell: str, text: str) -> None:
    ws[cell] = text
    ws[cell].font = Font(name="Calibri", bold=True, size=12)


def _write_headers(ws: Worksheet, row: int, headers: list[str]) -> None:
    for idx, h in enumerate(headers):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}{row}"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()


def _fy_cols() -> list[str]:
    """표 4·6: FY1..FY5 데이터 열 B..F."""
    return ["B", "C", "D", "E", "F"]


def _ov_col_for_fy(fy_idx: int) -> str:
    return OVERLAY_FY_COLS[fy_idx - 1]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_DASH)
    ws.column_dimensions["A"].width = 40
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "8. Dashboard — Word 심사보고서 복붙용"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:J1")

    # ---- Scenario meta scalars (rows 3–4) ----
    ws["A3"] = "Case"
    ws["B3"] = "=Case_Switch"
    c.define_name(wb, "DASH_Case", f"'{SHEET_DASH}'!$B$3")
    ws["A4"] = "Template Version"
    ws["B4"] = "v0.5"
    c.define_name(wb, "DASH_Version", f"'{SHEET_DASH}'!$B$4")

    # ---- 표 1. Valuation 요약 (title A6, headers row 7, data rows 8–10) ----
    _table_title(ws, "A6", "표 1. Valuation 요약")
    _write_headers(
        ws,
        7,
        ["방식", "Label", "Multiple", "EV (담보지분가치)", "Opco LTV", "누적 LTV"],
    )
    for i in (1, 2, 3):
        r = 7 + i
        ws[f"A{r}"] = f"방식 {i}"
        ws[f"B{r}"] = f"=DASH_Valuation_Method{i}_Label"
        ws[f"C{r}"] = f"=DASH_Valuation_Method{i}_Multiple"
        ws[f"D{r}"] = f"=DASH_Valuation_Method{i}_EV"
        ws[f"E{r}"] = f"=DASH_LTV_Method{i}_Opco"
        ws[f"F{r}"] = f"=DASH_LTV_Method{i}_Cumulative"
        for col in "CDEF":
            c.apply_key_output(ws[f"{col}{r}"])
        ws[f"C{r}"].number_format = c.NUM_FMT_MULTIPLE
        ws[f"D{r}"].number_format = c.NUM_FMT_ACCOUNTING
        ws[f"E{r}"].number_format = c.NUM_FMT_PERCENT
        ws[f"F{r}"].number_format = c.NUM_FMT_PERCENT

    # ---- 표 2. 이자지급가능성 요약 — 원래 형식(구분 + FY1..FY5 + Min, 영문 라벨) ----
    _table_title(ws, "A13", "표 2. 이자지급가능성 요약")
    _write_headers(ws, 14, ["구분", "FY1", "FY2", "FY3", "FY4", "FY5", "Min"])

    ws["A15"] = "Dividend Received"
    for i, col in enumerate("BCDEF"):
        ws[f"{col}15"] = f"=INDEX(Dividend_Row,{i + 1})"
        ws[f"{col}15"].number_format = c.NUM_FMT_ACCOUNTING
        c.define_name(wb, f"DASH_Div_FY{i + 1}", f"'{SHEET_DASH}'!${col}$15")

    ws["A16"] = "Holdco ICR"
    for i, col in enumerate("BCDEF"):
        ws[f"{col}16"] = f"=INDEX(Holdco_ICR_Row,{i + 1})"
        ws[f"{col}16"].number_format = c.NUM_FMT_MULTIPLE
    ws["G16"] = "=MIN(Holdco_ICR_Row)"
    ws["G16"].number_format = c.NUM_FMT_MULTIPLE
    c.apply_key_output(ws["G16"])
    c.define_name(wb, "DASH_ICR_Holdco_Min", f"'{SHEET_DASH}'!$G$16")

    ws["A17"] = "Opco ICR"
    for i, col in enumerate("BCDEF"):
        ws[f"{col}17"] = f"=INDEX(Opco_ICR_Row,{i + 1})"
        ws[f"{col}17"].number_format = c.NUM_FMT_MULTIPLE
    ws["G17"] = "=MIN(Opco_ICR_Row)"
    ws["G17"].number_format = c.NUM_FMT_MULTIPLE
    c.apply_key_output(ws["G17"])
    c.define_name(wb, "DASH_ICR_Opco_Min", f"'{SHEET_DASH}'!$G$17")

    ws["A18"] = "Opco DSCR"
    for i, col in enumerate("BCDEF"):
        ws[f"{col}18"] = f"=INDEX(Opco_DSCR_Row,{i + 1})"
        ws[f"{col}18"].number_format = c.NUM_FMT_MULTIPLE
    ws["G18"] = "=MIN(Opco_DSCR_Row)"
    ws["G18"].number_format = c.NUM_FMT_MULTIPLE
    c.apply_key_output(ws["G18"])
    c.define_name(wb, "DASH_DSCR_Min", f"'{SHEET_DASH}'!$G$18")

    ws["A20"] = "Net Leverage"
    for i, col in enumerate("BCDEF"):
        ws[f"{col}20"] = f"=INDEX(Net_Leverage_Row,{i + 1})"
        ws[f"{col}20"].number_format = c.NUM_FMT_MULTIPLE
        c.define_name(
            wb, f"DASH_Lev_NetLeverage_FY{i + 1}", f"'{SHEET_DASH}'!${col}$20"
        )

    # ---- 표 3. 만기상환가능성 요약 (title A22, data row 23) ----
    _table_title(ws, "A22", "표 3. 만기상환가능성 요약")
    ws["A23"] = "Exit Multiple (+ Active Δ)"
    ws["B23"] = "=DASH_Valuation_Method2_Multiple+Active_Exit_Multiple_Delta"
    ws["B23"].number_format = c.NUM_FMT_MULTIPLE
    c.apply_calc(ws["B23"])

    # ---- 표 4. 차주기준 자금수지표 (샘플 구조, FY1..FY5 = B..F) ----
    _table_title(ws, "A26", "표 4. 차주기준 자금수지표")
    hdr = [
        "구분",
        "'24.1H (인수시점)",
        "'25.1H",
        "'26.1H",
        "'27.1H",
        "'28.1H",
    ]
    _write_headers(ws, T4_HEADER_ROW, hdr)

    labels: list[tuple[str, str]] = [
        ("기초현금(a)", ""),
        ("대상주식 인수 관련 현금흐름", ""),
        ("  본건 인수금융 조달", ""),
        ("  재무대책(유상증자 등 조달)", ""),
        ("  대상회사 인수대금 지급", ""),
        ("(+)영업활동현금흐름(b)", ""),
        ("  EBITDA", ""),
        ("(-)투자활동현금흐름(c)", ""),
        ("  CAPEX", ""),
        ("(+)배당수익(d)", ""),
        ("원리금 상환재원(e=a+b+c+d)", ""),
        ("(+)추가 차입", ""),
        ("(-) 기존 차입금 원리금 상환", ""),
        ("   원금상환", ""),
        ("   이자비용 지급", ""),
        ("(-) 배당금 지급", ""),
        ("(-) 법인세 및 기타 현금흐름", ""),
        ("본건 인수금융 상환재원", ""),
        ("(-) 인수금융 이자비용", ""),
        ("(-) Tr.A 원금 상환 (Opco Senior)", ""),
        ("(-) Tr.B 원금 상환 (Opco 2L)", ""),
        ("(-) Tr.C 원금 상환 (Holdco)", ""),
        ("(-) 본건 Tr 원금 합계 (Tr.A+B+C)", ""),
        ("원리금 상환 이후 현금흐름", ""),
        ("리파이낸싱", ""),
        ("기말현금", ""),
    ]
    row_map = {
        "기초현금(a)": T4_ROW_기초,
        "대상주식 인수 관련 현금흐름": T4_ROW_인수관련,
        "  본건 인수금융 조달": T4_ROW_인수금융조달,
        "  재무대책(유상증자 등 조달)": T4_ROW_재무대책,
        "  대상회사 인수대금 지급": T4_ROW_인수대금지급,
        "(+)영업활동현금흐름(b)": T4_ROW_영업활동,
        "  EBITDA": T4_ROW_EBITDA,
        "(-)투자활동현금흐름(c)": T4_ROW_투자활동,
        "  CAPEX": T4_ROW_CAPEX,
        "(+)배당수익(d)": T4_ROW_배당수익,
        "원리금 상환재원(e=a+b+c+d)": T4_ROW_원리금상환재원,
        "(+)추가 차입": T4_ROW_추가차입,
        "(-) 기존 차입금 원리금 상환": T4_ROW_기존차입원리금,
        "   원금상환": T4_ROW_기존원금,
        "   이자비용 지급": T4_ROW_기존이자,
        "(-) 배당금 지급": T4_ROW_배당금지급,
        "(-) 법인세 및 기타 현금흐름": T4_ROW_법인세기타,
        "본건 인수금융 상환재원": T4_ROW_본건상환재원,
        "(-) 인수금융 이자비용": T4_ROW_본건이자,
        "(-) Tr.A 원금 상환 (Opco Senior)": T4_ROW_TrA,
        "(-) Tr.B 원금 상환 (Opco 2L)": T4_ROW_TrB,
        "(-) Tr.C 원금 상환 (Holdco)": T4_ROW_TrC,
        "(-) 본건 Tr 원금 합계 (Tr.A+B+C)": T4_ROW_Tr원금합계,
        "원리금 상환 이후 현금흐름": T4_ROW_원리금이후,
        "리파이낸싱": T4_ROW_리파이낸싱,
        "기말현금": T4_ROW_기말,
    }
    for label, _note in labels:
        r = row_map[label]
        ws[f"A{r}"] = label
        if label.startswith("  ") or label.startswith("   "):
            ws[f"A{r}"].alignment = Alignment(indent=1)

    fy_cols = _fy_cols()
    for fy_idx, col in enumerate(fy_cols, start=1):
        ov = _ov_col_for_fy(fy_idx)

        # 기초현금
        b = ws[f"{col}{T4_ROW_기초}"]
        if fy_idx == 1:
            b.value = 0
        else:
            prev = fy_cols[fy_idx - 2]
            b.value = f"={prev}{T4_ROW_기말}"
        c.apply_key_output(b)
        b.number_format = c.NUM_FMT_ACCOUNTING

        # 인수 관련
        ws[f"{col}{T4_ROW_인수금융조달}"].value = (
            f"=IF({fy_idx}=1,'{SHEET_INPUT}'!$B$9+'{SHEET_INPUT}'!$B$10+'{SHEET_INPUT}'!$B$11,0)"
        )
        ws[f"{col}{T4_ROW_재무대책}"].value = 0
        c.apply_input(ws[f"{col}{T4_ROW_재무대책}"])
        ws[f"{col}{T4_ROW_인수대금지급}"].value = (
            f"=IF({fy_idx}=1,-'{SHEET_INPUT}'!$B$7,0)"
        )
        ws[f"{col}{T4_ROW_인수관련}"].value = (
            f"={col}{T4_ROW_인수금융조달}+{col}{T4_ROW_재무대책}+{col}{T4_ROW_인수대금지급}"
        )

        # 영업 / EBITDA / 투자 / CAPEX / 배당
        ws[f"{col}{T4_ROW_EBITDA}"].value = (
            f"='{SHEET_OVERLAY}'!{ov}{STRESSED_EBITDA_ROW}"
        )
        ws[f"{col}{T4_ROW_영업활동}"].value = f"={col}{T4_ROW_EBITDA}"
        ws[f"{col}{T4_ROW_CAPEX}"].value = (
            f"=-'{SHEET_OVERLAY}'!{ov}{STRESSED_CAPEX_ROW}"
        )
        ws[f"{col}{T4_ROW_투자활동}"].value = f"={col}{T4_ROW_CAPEX}"
        ws[f"{col}{T4_ROW_배당수익}"].value = f"=INDEX(Dividend_Row,{fy_idx})"

        # 원리금 상환재원
        ws[f"{col}{T4_ROW_원리금상환재원}"].value = (
            f"={col}{T4_ROW_기초}+{col}{T4_ROW_인수관련}+{col}{T4_ROW_영업활동}"
            f"+{col}{T4_ROW_투자활동}+{col}{T4_ROW_배당수익}"
        )

        ws[f"{col}{T4_ROW_추가차입}"].value = 0
        c.apply_input(ws[f"{col}{T4_ROW_추가차입}"])

        ws[f"{col}{T4_ROW_기존원금}"].value = (
            f"=-(INDEX(Opco_Sr_Mand,{fy_idx})+INDEX(Opco_2L_Mand,{fy_idx}))"
        )
        ws[f"{col}{T4_ROW_기존이자}"].value = (
            f"=-(INDEX(Opco_Sr_Interest,{fy_idx})+INDEX(Opco_2L_Interest,{fy_idx}))"
        )
        ws[f"{col}{T4_ROW_기존차입원리금}"].value = (
            f"={col}{T4_ROW_기존원금}+{col}{T4_ROW_기존이자}"
        )

        ws[f"{col}{T4_ROW_배당금지급}"].value = 0
        c.apply_input(ws[f"{col}{T4_ROW_배당금지급}"])
        ws[f"{col}{T4_ROW_법인세기타}"].value = 0
        c.apply_input(ws[f"{col}{T4_ROW_법인세기타}"])

        ws[f"{col}{T4_ROW_본건상환재원}"].value = (
            f"={col}{T4_ROW_원리금상환재원}+{col}{T4_ROW_추가차입}+{col}{T4_ROW_기존차입원리금}"
            f"+{col}{T4_ROW_배당금지급}+{col}{T4_ROW_법인세기타}"
        )

        ws[f"{col}{T4_ROW_본건이자}"].value = (
            f"=-INDEX(Holdco_Interest,{fy_idx})"
        )

        deb_col = OVERLAY_FY_COLS[fy_idx - 1]
        ws[f"{col}{T4_ROW_TrA}"].value = f"=-INDEX(Opco_Sr_Mand,{fy_idx})"
        ws[f"{col}{T4_ROW_TrB}"].value = f"=-INDEX(Opco_2L_Mand,{fy_idx})"
        ws[f"{col}{T4_ROW_TrC}"].value = f"=-'{SHEET_DEBT}'!{deb_col}29"

        ws[f"{col}{T4_ROW_Tr원금합계}"].value = (
            f"={col}{T4_ROW_TrA}+{col}{T4_ROW_TrB}+{col}{T4_ROW_TrC}"
        )

        ws[f"{col}{T4_ROW_원리금이후}"].value = (
            f"={col}{T4_ROW_본건상환재원}+{col}{T4_ROW_본건이자}"
            f"+{col}{T4_ROW_TrA}+{col}{T4_ROW_TrB}+{col}{T4_ROW_TrC}"
        )

        ws[f"{col}{T4_ROW_리파이낸싱}"].value = 0
        c.apply_input(ws[f"{col}{T4_ROW_리파이낸싱}"])
        ws[f"{col}{T4_ROW_기말}"].value = (
            f"={col}{T4_ROW_원리금이후}+{col}{T4_ROW_리파이낸싱}"
        )

        for rr in (
            T4_ROW_인수관련,
            T4_ROW_인수금융조달,
            T4_ROW_인수대금지급,
            T4_ROW_영업활동,
            T4_ROW_EBITDA,
            T4_ROW_투자활동,
            T4_ROW_CAPEX,
            T4_ROW_배당수익,
            T4_ROW_원리금상환재원,
            T4_ROW_기존차입원리금,
            T4_ROW_기존원금,
            T4_ROW_기존이자,
            T4_ROW_본건상환재원,
            T4_ROW_본건이자,
            T4_ROW_TrA,
            T4_ROW_TrB,
            T4_ROW_TrC,
            T4_ROW_Tr원금합계,
            T4_ROW_원리금이후,
            T4_ROW_기말,
        ):
            cell = ws[f"{col}{rr}"]
            c.apply_key_output(cell)
            cell.number_format = c.NUM_FMT_ACCOUNTING

    # DASH_CFTable_* — 기존 8행×5열 Named Range (상세 표의 요약 행에 연결)
    for dash_idx, r in DASH_CF_SUMMARY_ROWS.items():
        c.define_name(wb, f"DASH_CFTable_Row{dash_idx}_Label", f"'{SHEET_DASH}'!$A${r}")
        for fy_idx, col in enumerate(fy_cols, start=1):
            c.define_name(
                wb,
                f"DASH_CFTable_Row{dash_idx}_FY{fy_idx}",
                f"'{SHEET_DASH}'!${col}${r}",
            )

    # ---- 표 5. 시나리오 메타 ----
    t5 = 55
    _table_title(ws, f"A{t5}", "표 5. 시나리오 메타")
    ws[f"A{t5 + 1}"] = "Revenue Growth Δ"
    ws[f"B{t5 + 1}"] = "=Active_Revenue_Growth_Delta"
    ws[f"B{t5 + 1}"].number_format = c.NUM_FMT_PERCENT
    ws[f"A{t5 + 2}"] = "EBITDA Margin Δ"
    ws[f"B{t5 + 2}"] = "=Active_EBITDA_Margin_Delta"
    ws[f"B{t5 + 2}"].number_format = c.NUM_FMT_PERCENT
    ws[f"A{t5 + 3}"] = "WACC Uplift (bp)"
    ws[f"B{t5 + 3}"] = "=Active_WACC_Uplift"
    ws[f"B{t5 + 3}"].number_format = c.NUM_FMT_BPS
    ws[f"A{t5 + 4}"] = "Exit Multiple Δ"
    ws[f"B{t5 + 4}"] = "=Active_Exit_Multiple_Delta"
    ws[f"B{t5 + 4}"].number_format = c.NUM_FMT_MULTIPLE

    irr_row = t5 + 6
    ws[f"A{irr_row}"] = "Sponsor IRR (v1.1+)"
    c.define_name(wb, "DASH_IRR_Sponsor", f"'{SHEET_DASH}'!$B${irr_row}")

    # ---- 표 6. 재무약정 준수여부 ----
    t6 = irr_row + 2
    _table_title(ws, f"A{t6}", "표 6. 재무약정 준수여부")
    ws.merge_cells(f"A{t6 + 1}:F{t6 + 1}")
    m = ws[f"A{t6 + 1}"]
    m.value = "[심사팀 스트레스 기준] 재무약정 준수 여부 (행 라벨은 수기로 바꿀 수 있음)"
    m.font = Font(name="Calibri", bold=True, size=10)
    m.alignment = Alignment(wrap_text=True)

    cov_hdr = ["구분", "'24년", "'25년", "'26년", "'27년", "'28년"]
    _write_headers(ws, t6 + 2, cov_hdr)
    base_r = t6 + 3

    ws[f"G{t6 + 1}"] = "약정 항목 선택"
    ws[f"G{t6 + 1}"].font = Font(name="Calibri", bold=True, size=9)
    dv_item = DataValidation(
        type="list",
        formula1='"차주 부채비율(%),차주 이자보상배율(x),차주 Net Debt/EBITDA,직접입력"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_item)

    # ① 부채비율 — 수기 입력 (BS 연결 시 외부에서 붙여넣기). A열 라벨은 필요 시 직접 수정.
    ws[f"A{base_r}"] = "① 차주 부채비율(%)"
    dv_item.add(f"G{base_r}")
    for col in "BCDEF":
        ws[f"{col}{base_r}"] = None
        c.apply_input(ws[f"{col}{base_r}"])
    ws[f"A{base_r + 1}"] = "재무약정 준수기준"
    ws[f"B{base_r + 1}"] = "200% 이하 유지"
    ws.merge_cells(f"B{base_r + 1}:F{base_r + 1}")
    ws[f"A{base_r + 2}"] = "준수 여부"
    for col in "BCDEF":
        ws[f"{col}{base_r + 2}"] = ""
        c.apply_input(ws[f"{col}{base_r + 2}"])

    # ② 이자보상배율
    r2 = base_r + 3
    ws[f"A{r2}"] = "② 차주 이자보상배율(배)"
    dv_item.add(f"G{r2}")
    for col in "BCDEF":
        ws[f"{col}{r2}"] = f"=INDEX(Opco_ICR_Row,{list('BCDEF').index(col) + 1})"
        c.apply_key_output(ws[f"{col}{r2}"])
        ws[f"{col}{r2}"].number_format = c.NUM_FMT_MULTIPLE
    ws[f"A{r2 + 1}"] = "재무약정 준수기준"
    ws[f"B{r2 + 1}"] = "1.2x (FY별 수기)"
    ws.merge_cells(f"B{r2 + 1}:F{r2 + 1}")
    ws[f"A{r2 + 2}"] = "준수 여부"
    for col in "BCDEF":
        ws[f"{col}{r2 + 2}"] = ""
        c.apply_input(ws[f"{col}{r2 + 2}"])

    # ③ Net Debt/EBITDA
    r3 = r2 + 3
    ws[f"A{r3}"] = "③ 차주 연결기준 Net Debt/EBITDA"
    dv_item.add(f"G{r3}")
    for i, col in enumerate("BCDEF"):
        ws[f"{col}{r3}"] = f"=INDEX(Net_Leverage_Row,{i + 1})"
        c.apply_key_output(ws[f"{col}{r3}"])
        ws[f"{col}{r3}"].number_format = c.NUM_FMT_MULTIPLE
    ws[f"A{r3 + 1}"] = "재무약정 준수기준"
    ws[f"B{r3 + 1}"] = ""
    c.apply_input(ws[f"B{r3 + 1}"])
    ws.merge_cells(f"B{r3 + 1}:F{r3 + 1}")
    ws[f"A{r3 + 2}"] = "준수 여부"
    for col in "BCDEF":
        ws[f"{col}{r3 + 2}"] = ""
        c.apply_input(ws[f"{col}{r3 + 2}"])

    # ---- 표 7. 만기시점 LTV 분석 (만기=FY5, 평가방식 2 기준) ----
    t7 = r3 + 4
    _table_title(ws, f"A{t7}", "표 7. 만기시점 LTV 분석")
    r_e, r_m, r_ev, r_nd, r_eq, r_rf, r_ltv = range(t7 + 1, t7 + 8)

    ws[f"A{r_e}"] = "만기시점 EBITDA"
    ws[f"B{r_e}"] = f"='{SHEET_OVERLAY}'!$I${STRESSED_EBITDA_ROW}"
    ws[f"A{r_m}"] = "적용 EV/EBITDA 멀티플"
    ws[f"B{r_m}"] = "=DASH_Valuation_Method2_Multiple"
    ws[f"A{r_ev}"] = "기업가치"
    ws[f"B{r_ev}"] = f"=B{r_e}*B{r_m}"
    ws[f"A{r_nd}"] = "만기시점 Net Debt"
    ws[f"B{r_nd}"] = (
        f"='{SHEET_DEBT}'!$I$11+'{SHEET_DEBT}'!$I$21+'{SHEET_DEBT}'!$I$31"
    )
    ws[f"A{r_eq}"] = "담보지분가치(지분율)"
    ws[f"B{r_eq}"] = f"='{SHEET_RETURNS}'!$F$12"
    ws[f"A{r_rf}"] = "리파이낸싱 필요액 (Tr.A+Tr.B+RCF)"
    ws[f"C{r_rf}"] = "RCF 등 가산"
    ws[f"D{r_rf}"] = 0
    c.apply_input(ws[f"D{r_rf}"])
    ws[f"D{r_rf}"].number_format = c.NUM_FMT_ACCOUNTING
    ws[f"B{r_rf}"].value = f"='{SHEET_DEBT}'!$I$11+'{SHEET_DEBT}'!$I$21+D{r_rf}"

    ws[f"A{r_ltv}"] = "만기시점 인수금융 누적 LTV"
    ws[f"B{r_ltv}"] = f"=IFERROR(B{r_nd}/B{r_eq},\"\")"

    for rr, fmt in (
        (r_e, c.NUM_FMT_ACCOUNTING),
        (r_m, c.NUM_FMT_MULTIPLE),
        (r_ev, c.NUM_FMT_ACCOUNTING),
        (r_nd, c.NUM_FMT_ACCOUNTING),
        (r_eq, c.NUM_FMT_ACCOUNTING),
        (r_rf, c.NUM_FMT_ACCOUNTING),
        (r_ltv, c.NUM_FMT_PERCENT),
    ):
        cell = ws[f"B{rr}"]
        cell.number_format = fmt
        c.apply_key_output(cell)

    return ws
