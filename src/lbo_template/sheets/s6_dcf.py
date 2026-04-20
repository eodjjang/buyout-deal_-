"""6_DCF_Valuation — FCFF + mid-year discount + Gordon TV (design §6)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName

from lbo_template.layout import SHEET_DCF, SHEET_INPUT, SHEET_OVERLAY
from lbo_template import conventions as c
from lbo_template.sheets.s3_overlay import (
    OVERLAY_CASH_TAXES_ROW,
    STRESSED_CAPEX_ROW,
    STRESSED_EBITDA_ROW,
    STRESSED_NWC_ROW,
)

# E..I = FY1..FY5, J = TV
DCF_COLS = ["E", "F", "G", "H", "I"]
TV_COL = "J"
PERIODS = {"E": 0.5, "F": 1.5, "G": 2.5, "H": 3.5, "I": 4.5}

ROWS_FY: list[tuple[str, str | None]] = [
    ("Stressed EBITDA", "='{ov}'!{c}{ebitda_row}"),
    # Overlay row 18 = Cash Taxes (플랜 스니펫 row 20은 현재 그리드와 불일치)
    ("(-) Cash Taxes on EBIT", "='{ov}'!{c}{cash_tax_row}"),
    ("(-) Capex", "='{ov}'!{c}{capex_row}"),
    ("(-) ΔNWC", "='{ov}'!{c}{nwc_row}"),
    ("FCFF", "={c}5-{c}6-{c}7-{c}8"),
    ("WACC", "=Base_WACC+Active_WACC_Uplift/10000"),
    ("Discount Period", None),
    ("Discount Factor", "=1/(1+{c}10)^{c}11"),
    ("PV of FCFF", "={c}9*{c}12"),
]

TV_ROWS = [
    ("Terminal Value (Gordon)", "=I9*(1+Perm_Growth)/(I10-Perm_Growth)"),
    ("PV of TV", "=J14/(1+I10)^5.0"),
]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_DCF)
    ws.column_dimensions["A"].width = 36
    for col in DCF_COLS + [TV_COL]:
        ws.column_dimensions[col].width = 14

    ws["A1"] = "6. DCF Valuation (Stressed)"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:J1")

    fy_labels = ["FY1", "FY2", "FY3", "FY4", "FY5", "TV"]
    for col, label in zip(DCF_COLS + [TV_COL], fy_labels):
        hdr = ws[f"{col}3"]
        hdr.value = label
        hdr.font = c.column_header_font()
        hdr.fill = c.column_header_fill()

    fmt_ctx = {
        "ov": SHEET_OVERLAY,
        "inp": SHEET_INPUT,
        "ebitda_row": STRESSED_EBITDA_ROW,
        "capex_row": STRESSED_CAPEX_ROW,
        "nwc_row": STRESSED_NWC_ROW,
        "cash_tax_row": OVERLAY_CASH_TAXES_ROW,
    }

    for idx, (label, template) in enumerate(ROWS_FY):
        r = 5 + idx
        ws.cell(row=r, column=1, value=label)
        for col in DCF_COLS:
            cell = ws[f"{col}{r}"]
            if label == "Discount Period":
                cell.value = PERIODS[col]
                c.apply_calc(cell)
                cell.number_format = "0.0"
            else:
                assert template is not None
                cell.value = template.format(c=col, **fmt_ctx)
                c.apply_calc(cell)
                if label == "WACC":
                    cell.number_format = c.NUM_FMT_PERCENT
                elif label == "Discount Factor":
                    cell.number_format = "0.0000"
                else:
                    cell.number_format = c.NUM_FMT_ACCOUNTING
                if label in ("FCFF", "PV of FCFF"):
                    c.apply_key_output(cell)

    j11 = ws[f"{TV_COL}11"]
    j11.value = 5.0
    j11.number_format = "0.0"
    c.apply_calc(j11)

    ws.cell(row=14, column=1, value="Terminal Value (Gordon)")
    j14 = ws[f"{TV_COL}14"]
    j14.value = TV_ROWS[0][1]
    c.apply_key_output(j14)
    j14.number_format = c.NUM_FMT_ACCOUNTING

    ws.cell(row=15, column=1, value="PV of TV")
    j15 = ws[f"{TV_COL}15"]
    j15.value = TV_ROWS[1][1]
    c.apply_key_output(j15)
    j15.number_format = c.NUM_FMT_ACCOUNTING

    ws.cell(row=17, column=1, value="EV (PV 합계)")
    e17 = ws["E17"]
    e17.value = "=SUM(E13:I13)+J15"
    c.apply_key_output(e17)
    e17.number_format = c.NUM_FMT_ACCOUNTING

    ws.cell(row=18, column=1, value="(+) 비영업자산")
    e18 = ws["E18"]
    e18.value = 0
    c.apply_input(e18)
    e18.number_format = c.NUM_FMT_ACCOUNTING

    ws.cell(row=19, column=1, value="(-) Net Debt (Closing)")
    e19 = ws["E19"]
    e19.value = f"='{SHEET_INPUT}'!B6"
    c.apply_calc(e19)
    e19.font = c.crosstab_link_font()
    e19.number_format = c.NUM_FMT_ACCOUNTING

    ws.cell(row=20, column=1, value="= 담보기준 Equity Value")
    e20 = ws["E20"]
    e20.value = "=E17+E18-E19"
    c.apply_key_output(e20)
    e20.number_format = c.NUM_FMT_ACCOUNTING

    ws.cell(row=23, column=1, value="Base WACC (사용자 입력)")
    b23 = ws["B23"]
    b23.value = 0.10
    c.apply_input(b23)
    b23.number_format = c.NUM_FMT_PERCENT
    wb.defined_names["Base_WACC"] = DefinedName(
        "Base_WACC", attr_text=f"'{SHEET_DCF}'!$B$23"
    )

    wb.defined_names["DCF_EV"] = DefinedName(
        "DCF_EV", attr_text=f"'{SHEET_DCF}'!$E$17"
    )
    wb.defined_names["DCF_Equity_Value"] = DefinedName(
        "DCF_Equity_Value", attr_text=f"'{SHEET_DCF}'!$E$20"
    )

    return ws
