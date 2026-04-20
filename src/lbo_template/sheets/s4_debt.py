"""4_Debt_Schedule — Opco Senior / Opco 2nd Lien / Holdco Sub (design §4)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment

from lbo_template.layout import SHEET_DEBT
from lbo_template import conventions as c

_SECTION_TITLE_FONT = Font(name="Calibri", bold=True, size=11)

# Tranche block: section title row + 6 metric rows (Opening … Ending).
TRANCHE_ROWS = [
    "Opening Balance",
    "Interest Rate",
    "Interest Expense",
    "Mandatory Amortization",
    "Cash Sweep Applied",
    "Ending Balance",
]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_DEBT)
    ws.column_dimensions["A"].width = 36
    for col in c.FY_AXIS_COLUMNS:
        ws.column_dimensions[col].width = 14

    ws["A1"] = "4. Debt Schedule"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:I1")

    for col, label in zip(c.FY_AXIS_COLUMNS, c.FY_AXIS_LABELS):
        cell = ws[f"{col}3"]
        cell.value = label
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    tranches: list[tuple[str, str, int, str]] = [
        ("Opco Senior TL", "Opco_Senior_Principal", 5, "SENIOR"),
        ("Opco 2nd Lien", "Opco_2L_Principal", 15, "SECOND"),
        ("Holdco Sub", "Holdco_Sub_Principal", 25, "HOLDCO"),
    ]

    for tranche_name, principal_named, start_row, tag in tranches:
        hdr = ws.cell(row=start_row, column=1, value=tranche_name)
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=9)
        c.apply_section_header(hdr)

        for idx, label in enumerate(TRANCHE_ROWS):
            r = start_row + 1 + idx
            ws.cell(row=r, column=1, value=label)
            for col_idx, col in enumerate(c.FY_AXIS_COLUMNS):
                cell = ws[f"{col}{r}"]
                if col_idx < 3:
                    cell.value = None
                    continue

                if label == "Opening Balance":
                    if col == "E":
                        cell.value = f"={principal_named}"
                    else:
                        prev = c.FY_AXIS_COLUMNS[col_idx - 1]
                        cell.value = f"={prev}{r + 5}"
                    c.apply_calc(cell)
                elif label == "Interest Rate":
                    if col == "E":
                        cell.value = None
                        c.apply_input(cell)
                        cell.number_format = c.NUM_FMT_PERCENT
                    else:
                        prev = c.FY_AXIS_COLUMNS[col_idx - 1]
                        cell.value = f"={prev}{r}"
                        c.apply_calc(cell)
                        cell.number_format = c.NUM_FMT_PERCENT
                elif label == "Interest Expense":
                    cell.value = f"={col}{r - 2}*{col}{r - 1}"
                    c.apply_calc(cell)
                elif label == "Mandatory Amortization":
                    cell.value = 0
                    c.apply_input(cell)
                elif label == "Cash Sweep Applied":
                    if tag == "HOLDCO":
                        cell.value = 0
                        c.apply_calc(cell)
                    elif tag == "SENIOR":
                        cell.value = (
                            f"=MIN(MAX(0,{col}{r - 3}-{col}{r - 1}),"
                            f"IFERROR(Opco_Sweep_Avail_{col},0))"
                        )
                        c.apply_calc(cell)
                    elif tag == "SECOND":
                        cell.value = (
                            f"=MIN(MAX(0,{col}{r - 3}-{col}{r - 1}),"
                            f"IFERROR(Opco_Sweep_Avail_{col}-{col}{r - 10},0))"
                        )
                        c.apply_calc(cell)
                elif label == "Ending Balance":
                    if tag == "HOLDCO":
                        cell.value = (
                            f"=IF(Holdco_PIK_Mode=\"PIK\","
                            f"MAX(0,{col}{r - 5}+{col}{r - 3}-{col}{r - 2}-{col}{r - 1}),"
                            f"MAX(0,{col}{r - 5}-{col}{r - 2}-{col}{r - 1}))"
                        )
                    else:
                        cell.value = f"=MAX(0,{col}{r - 5}-{col}{r - 2}-{col}{r - 1})"
                    c.apply_key_output(cell)

    ws["A35"] = "Holdco 이자 지급방식"
    ws["A35"].font = _SECTION_TITLE_FONT
    ws["A35"].alignment = Alignment(horizontal="left", vertical="center")
    b35 = ws["B35"]
    b35.value = "Cash"
    c.apply_input(b35)
    dv_pik = DataValidation(type="list", formula1='"Cash,PIK"', allow_blank=False)
    ws.add_data_validation(dv_pik)
    dv_pik.add("B35")
    wb.defined_names["Holdco_PIK_Mode"] = DefinedName(
        "Holdco_PIK_Mode", attr_text=f"'{SHEET_DEBT}'!$B$35"
    )

    ws["A37"] = "Cash Sweep % (Opco excess cash → Senior prepay)"
    ws["A37"].font = _SECTION_TITLE_FONT
    ws["A37"].alignment = Alignment(horizontal="left", vertical="center")
    b37 = ws["B37"]
    b37.value = 1.0
    c.apply_input(b37)
    b37.number_format = c.NUM_FMT_PERCENT
    wb.defined_names["Sweep_Pct"] = DefinedName(
        "Sweep_Pct", attr_text=f"'{SHEET_DEBT}'!$B$37"
    )

    wb.defined_names["Opco_Sr_Interest"] = DefinedName(
        "Opco_Sr_Interest", attr_text=f"'{SHEET_DEBT}'!$E$8:$I$8"
    )
    wb.defined_names["Opco_2L_Interest"] = DefinedName(
        "Opco_2L_Interest", attr_text=f"'{SHEET_DEBT}'!$E$18:$I$18"
    )
    wb.defined_names["Holdco_Interest"] = DefinedName(
        "Holdco_Interest", attr_text=f"'{SHEET_DEBT}'!$E$28:$I$28"
    )
    wb.defined_names["Opco_Sr_Mand"] = DefinedName(
        "Opco_Sr_Mand", attr_text=f"'{SHEET_DEBT}'!$E$9:$I$9"
    )
    wb.defined_names["Opco_2L_Mand"] = DefinedName(
        "Opco_2L_Mand", attr_text=f"'{SHEET_DEBT}'!$E$19:$I$19"
    )

    return ws
