"""9_Peer_Summary — unified trading + transaction aggregation with Include toggles."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from lbo_template import conventions as c
from lbo_template.layout import SHEET_9A, SHEET_9B, SHEET_9C, SHEET_PEER


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_PEER)
    ws.column_dimensions["A"].width = 24
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 14

    ws["A1"] = "9. Peer Summary — 통합 집계 (Trading + Transaction)"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:K1")

    # --- Trading Peer Summary (rows 3..20) ---
    ws["A3"] = "Trading Peer Summary"
    ws["A3"].font = Font(name="Calibri", bold=True, size=12)

    trading_headers = [
        "Peer Name",
        "Source",
        "EV/LTM EBITDA",
        "EV/FY-1",
        "EV/FY-2",
        "EV/NTM",
        "PBR",
        "Net Debt/LTM EBITDA",
        "Include ✓",
        "Memo",
    ]
    for idx, h in enumerate(trading_headers):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}5"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    for r in range(6, 21):
        peer_r = r - 3
        ws[f"A{r}"] = f"='{SHEET_9A}'!A{peer_r}"
        ws[f"B{r}"] = "CIQ Trading"
        ws[f"C{r}"] = f"='{SHEET_9A}'!J{peer_r}"
        ws[f"D{r}"] = f"='{SHEET_9A}'!K{peer_r}"
        ws[f"E{r}"] = f"='{SHEET_9A}'!L{peer_r}"
        ws[f"F{r}"] = f"='{SHEET_9A}'!M{peer_r}"
        ws[f"G{r}"] = ""
        ws[f"H{r}"] = f"='{SHEET_9A}'!N{peer_r}"
        ws[f"I{r}"] = True
        c.apply_input(ws[f"I{r}"])
        for col in "ABCDEFGH":
            ws[f"{col}{r}"].font = c.crosstab_link_font()

    ws["A22"] = "Mean (Included only)"
    ws["A23"] = "Median (Included only)"
    ws["A24"] = "Min / Max"
    ws["A27"] = "3개년 평균의 평균 (Applied Trading Multiple)"

    for col in "CDEF":
        ws[f"{col}22"] = f'=IFERROR(AVERAGEIF($I$6:$I$20,TRUE,{col}$6:{col}$20),"")'
        ws[f"{col}23"] = f'=IFERROR(MEDIAN(IF($I$6:$I$20=TRUE,{col}$6:{col}$20)),"")'
        ws[f"{col}24"] = f'=MIN({col}$6:{col}$20)&" / "&MAX({col}$6:{col}$20)'
        c.apply_calc(ws[f"{col}22"])
        c.apply_calc(ws[f"{col}23"])
        c.apply_calc(ws[f"{col}24"])
        ws[f"{col}22"].number_format = c.NUM_FMT_MULTIPLE
        ws[f"{col}23"].number_format = c.NUM_FMT_MULTIPLE

    ws["C27"] = "=(C22+D22+E22)/3"
    c.apply_key_output(ws["C27"])
    ws["C27"].number_format = c.NUM_FMT_MULTIPLE
    c.define_name(wb, "Applied_Trading_Multiple", f"'{SHEET_PEER}'!$C$27")

    ws["G27"] = '=IFERROR(AVERAGEIF($I$6:$I$20,TRUE,G$6:G$20),"")'
    c.apply_key_output(ws["G27"])
    ws["G27"].number_format = c.NUM_FMT_MULTIPLE
    c.define_name(wb, "Applied_Trading_PBR", f"'{SHEET_PEER}'!$G$27")

    # --- Transaction Comps Summary (rows 30..65) ---
    ws["A30"] = "Transaction Comps Summary"
    ws["A30"].font = Font(name="Calibri", bold=True, size=12)

    tx_headers = [
        "Transaction ID",
        "Source",
        "Announced",
        "Target",
        "Buyer Type",
        "EV/LTM EBITDA",
        "EV/LTM Rev",
        "Deal Value Disclosed?",
        "Include ✓",
        "Memo",
    ]
    for idx, h in enumerate(tx_headers):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}32"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    for r in range(33, 53):
        src_r = r - 30
        ws[f"A{r}"] = f"='{SHEET_9B}'!A{src_r}"
        ws[f"B{r}"] = "CIQ M&A"
        ws[f"C{r}"] = f"='{SHEET_9B}'!B{src_r}"
        ws[f"D{r}"] = f"='{SHEET_9B}'!D{src_r}"
        ws[f"E{r}"] = f"='{SHEET_9B}'!H{src_r}"
        ws[f"F{r}"] = f"='{SHEET_9B}'!N{src_r}"
        ws[f"G{r}"] = f"='{SHEET_9B}'!M{src_r}"
        ws[f"H{r}"] = f'=IF(ISNUMBER(F{r}),"Yes","No")'
        ws[f"I{r}"] = f'=AND(H{r}="Yes",E{r}<>"Strategic")'
        c.apply_calc(ws[f"I{r}"])
        for col in "ABCDEFGH":
            ws[f"{col}{r}"].font = c.crosstab_link_font()

    for r in range(53, 63):
        src_r = r - 50
        ws[f"A{r}"] = f"='{SHEET_9C}'!A{src_r}"
        ws[f"B{r}"] = f"='{SHEET_9C}'!P{src_r}"
        ws[f"D{r}"] = f"='{SHEET_9C}'!D{src_r}"
        ws[f"F{r}"] = f"='{SHEET_9C}'!N{src_r}"
        ws[f"I{r}"] = f"='{SHEET_9C}'!R{src_r}"
        c.apply_calc(ws[f"A{r}"])
        c.apply_calc(ws[f"B{r}"])
        c.apply_calc(ws[f"D{r}"])
        c.apply_calc(ws[f"F{r}"])
        c.apply_calc(ws[f"I{r}"])
        for col in "ABCDEFGH":
            ws[f"{col}{r}"].font = c.crosstab_link_font()

    ws["A65"] = "Mean (Included only)"
    ws["A66"] = "Median"
    ws["A67"] = "Trimmed Mean (상하 10% 제외, Applied Transaction Multiple)"

    ws["F65"] = '=IFERROR(AVERAGEIF($I$33:$I$62,TRUE,F$33:F$62),"")'
    ws["F66"] = '=IFERROR(MEDIAN(IF($I$33:$I$62=TRUE,F$33:F$62)),"")'
    ws["F67"] = '=IFERROR(TRIMMEAN(IF($I$33:$I$62=TRUE,F$33:F$62),0.2),"")'
    c.apply_calc(ws["F65"])
    c.apply_calc(ws["F66"])
    c.apply_key_output(ws["F67"])
    ws["F65"].number_format = c.NUM_FMT_MULTIPLE
    ws["F66"].number_format = c.NUM_FMT_MULTIPLE
    ws["F67"].number_format = c.NUM_FMT_MULTIPLE
    c.define_name(wb, "Applied_Transaction_Multiple", f"'{SHEET_PEER}'!$F$67")

    return ws
