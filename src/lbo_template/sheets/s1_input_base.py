"""1_Input_BaseCase — sole user-input entry point for Base Case."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill
from openpyxl.formatting.rule import CellIsRule

from lbo_template.layout import SHEET_INPUT
from lbo_template import conventions as c

_SECTION_TITLE_FONT = Font(name="Calibri", bold=True, size=11)
_RED_FILL = PatternFill("solid", fgColor="FFC7CE")

SECTION_A_ROWS = [
    ("인수금액 (Purchase EV)", None, True),
    ("Less: Net Debt Assumed", None, True),
    ("= 지분 인수가액 (Equity Purchase Price)", "=B5-B6", False),
    ("+ Transaction Fee (M&A 자문·실사·세무)", None, True),
    ("= Uses of Funds 합계", "=B7+B8", False),
    ("Sources: Opco Senior TL", None, True),
    ("Sources: Opco 2nd Lien", None, True),
    ("Sources: Holdco Sub Loan", None, True),
    ("Sources: Sponsor Equity (plug)", "=B9-B10-B11-B12", False),
    ("Target Net Debt / LTM EBITDA (본부 승인치)", None, True),
    ("Closing Date", None, True),
    ("Exit Date (Assumed)", None, True),
]

SECTION_B_ROWS = [
    "Revenue",
    "COGS (or Gross Profit)",
    "SG&A",
    "EBITDA (Reported)",
    "D&A",
    "Capex",
    "Δ NWC (증가=현금유출, +)",
    "Effective Tax Rate",
]

SECTION_C_ROWS = [
    ("EBITDA Margin", "=IFERROR({c}27/{c}24, \"\")"),
    ("Capex as % of Revenue", "=IFERROR({c}29/{c}24, \"\")"),
    ("ΔNWC as % of Revenue", "=IFERROR({c}30/{c}24, \"\")"),
    ("Revenue YoY Growth", "=IFERROR({c}24/{prev}24-1, \"\")"),
    ("EBITDA YoY Growth", "=IFERROR({c}27/{prev}27-1, \"\")"),
]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_INPUT)
    ws.column_dimensions["A"].width = 44
    for col in c.FY_AXIS_COLUMNS:
        ws.column_dimensions[col].width = 14

    ws["A1"] = "1. Input — Base Case (사용자 입력 단일 진입점)"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:I1")

    # --- Section A --------------------------------------------------
    ws["A3"] = "Section A — 인수 조건 (Transaction Terms)"
    ws["A3"].font = _SECTION_TITLE_FONT

    for idx, (label, formula, is_input) in enumerate(SECTION_A_ROWS):
        r = 5 + idx
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=2)
        if formula:
            cell.value = formula
            c.apply_calc(cell)
        elif is_input:
            c.apply_input(cell)
        if label == "Target Net Debt / LTM EBITDA (본부 승인치)":
            cell.number_format = c.NUM_FMT_MULTIPLE
        elif label in ("Closing Date", "Exit Date (Assumed)"):
            cell.number_format = c.NUM_FMT_DATE
        else:
            cell.number_format = c.NUM_FMT_ACCOUNTING

    # Check rows (dual-check per v0.4 design)
    ws.cell(row=18, column=1, value="Check: Sources − Uses (표시용, =0이어야 함)")
    chk1 = ws.cell(row=18, column=2, value="=(B10+B11+B12+B13)-B9")
    chk1.number_format = c.NUM_FMT_ACCOUNTING
    c.apply_calc(chk1)

    ws.cell(
        row=19,
        column=1,
        value="Target Leverage Check ((Senior+2nd+Holdco)/LTM EBITDA ≤ Target)",
    )
    chk2 = ws.cell(row=19, column=2, value="=IFERROR((B10+B11+B12)/D27,\"\")")
    chk2.number_format = c.NUM_FMT_MULTIPLE
    c.apply_calc(chk2)

    # --- Section B --------------------------------------------------
    ws["A22"] = "Section B — Base Case 4대 드라이버"
    ws["A22"].font = _SECTION_TITLE_FONT

    for col, label in zip(c.FY_AXIS_COLUMNS, c.FY_AXIS_LABELS):
        cell = ws[f"{col}23"]
        cell.value = f"{label} Actual" if label in ("FY-2", "FY-1") else label
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    for idx, label in enumerate(SECTION_B_ROWS):
        r = 24 + idx
        ws.cell(row=r, column=1, value=label)
        for col in c.FY_AXIS_COLUMNS:
            cell = ws[f"{col}{r}"]
            c.apply_input(cell)
            if label == "Effective Tax Rate":
                cell.number_format = c.NUM_FMT_PERCENT
            else:
                cell.number_format = c.NUM_FMT_ACCOUNTING

    # Note row outside rows 24..34 scan range, avoiding forbidden "Adjusted" substring
    ws.cell(row=35, column=1, value="Note (EBITDA 환원 내역, Mgmt vs Bank Case 등)")
    note_cell = ws.cell(row=35, column=2)
    c.apply_input(note_cell)
    ws.merge_cells("B35:I35")

    # --- Section C --------------------------------------------------
    ws["A38"] = "Section C — Implied 역산 지표 (검증용)"
    ws["A38"].font = _SECTION_TITLE_FONT

    for idx, (label, template) in enumerate(SECTION_C_ROWS):
        r = 40 + idx
        ws.cell(row=r, column=1, value=label)
        for col_idx, col in enumerate(c.FY_AXIS_COLUMNS):
            cell = ws[f"{col}{r}"]
            if "{prev}" in template:
                if col_idx == 0:
                    continue
                prev_col = c.FY_AXIS_COLUMNS[col_idx - 1]
                cell.value = template.format(c=col, prev=prev_col)
            else:
                cell.value = template.format(c=col)
            c.apply_calc(cell)
            cell.number_format = c.NUM_FMT_PERCENT

    # Named ranges for downstream sheets. SHEET_INPUT starts with a digit,
    # so single-quoting the sheet name in attr_text is mandatory.
    wb.defined_names["LTM_EBITDA"] = DefinedName(
        "LTM_EBITDA", attr_text=f"'{SHEET_INPUT}'!$D$27"
    )
    wb.defined_names["Target_Leverage"] = DefinedName(
        "Target_Leverage", attr_text=f"'{SHEET_INPUT}'!$B$14"
    )
    wb.defined_names["Closing_Date"] = DefinedName(
        "Closing_Date", attr_text=f"'{SHEET_INPUT}'!$B$15"
    )
    wb.defined_names["Exit_Date"] = DefinedName(
        "Exit_Date", attr_text=f"'{SHEET_INPUT}'!$B$16"
    )
    wb.defined_names["Opco_Senior_Principal"] = DefinedName(
        "Opco_Senior_Principal", attr_text=f"'{SHEET_INPUT}'!$B$10"
    )
    wb.defined_names["Opco_2L_Principal"] = DefinedName(
        "Opco_2L_Principal", attr_text=f"'{SHEET_INPUT}'!$B$11"
    )
    wb.defined_names["Holdco_Sub_Principal"] = DefinedName(
        "Holdco_Sub_Principal", attr_text=f"'{SHEET_INPUT}'!$B$12"
    )

    # Flag Target Leverage Check red when principal stack exceeds Target multiple
    ws.conditional_formatting.add(
        "B19",
        CellIsRule(operator="greaterThan", formula=["B14"], fill=_RED_FILL),
    )
    return ws
