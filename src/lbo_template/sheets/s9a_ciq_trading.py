"""9a_CIQ_Trading_Raw — Plug-in primary, Paste fallback."""
from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from lbo_template import conventions as c
from lbo_template.layout import SHEET_9A

HEADERS = [
    "Company Name",
    "CIQ ID / Ticker",
    "Country",
    "Currency",
    "Market Cap",
    "Enterprise Value",
    "LTM Revenue",
    "LTM EBITDA",
    "LTM EBITDA Margin %",
    "EV / LTM EBITDA",
    "EV / FY-1 EBITDA",
    "EV / FY-2 EBITDA",
    "EV / NTM EBITDA",
    "Net Debt / LTM EBITDA",
    "LTM Period End Date",
]

# CIQ per column (B{r} = Ticker/식별자). 기간은 문자열 "LTM"이 아니라 플러그인 상수
# IQ_LTM / IQ_NTM / IQ_FY ± n — S&P Capital IQ Excel Plug-in Manual.
CIQ_FORMULAS = {
    "A": '=IFERROR(CIQ($B{r},"IQ_COMPANY_NAME"),"")',
    "C": '=IFERROR(CIQ($B{r},"IQ_COUNTRY_NAME"),"")',
    "D": '=IFERROR(CIQ($B{r},"IQ_TRADING_CURRENCY"),"")',
    "E": '=IFERROR(CIQ($B{r},"IQ_MARKETCAP"),"")',
    "F": '=IFERROR(CIQ($B{r},"IQ_TEV"),"")',
    "G": '=IFERROR(CIQ($B{r},"IQ_TOTAL_REV",IQ_LTM),"")',
    "H": '=IFERROR(CIQ($B{r},"IQ_EBITDA",IQ_LTM),"")',
    "I": '=IFERROR(CIQ($B{r},"IQ_EBITDA_MARGIN",IQ_LTM),"")',
    "J": '=IFERROR(CIQ($B{r},"IQ_TEV_EBITDA",IQ_LTM),"")',
    "K": '=IFERROR(CIQ($B{r},"IQ_TEV_EBITDA",IQ_FY - 1),"")',
    "L": '=IFERROR(CIQ($B{r},"IQ_TEV_EBITDA",IQ_FY - 2),"")',
    "M": '=IFERROR(CIQ($B{r},"IQ_TEV_EBITDA",IQ_NTM),"")',
    "N": '=IFERROR(CIQ($B{r},"IQ_NET_DEBT_EBITDA",IQ_LTM),"")',
    "O": '=IFERROR(CIQ($B{r},"IQ_PERIODDATE",IQ_LTM),"")',
}


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_9A)
    ws.column_dimensions["A"].width = 20
    for col_idx, _h in enumerate(HEADERS):
        col = chr(ord("A") + col_idx)
        if col != "A":
            ws.column_dimensions[col].width = 14

    # Mode cell (row 1)
    ws["A1"] = "Mode"
    ws["A1"].font = Font(name="Calibri", bold=True)
    ws["B1"] = '=IF(ISFORMULA(C3),"Plug-in","⚠ Paste Fallback — 마스터 재배포 필요")'
    c.apply_calc(ws["B1"])

    # Last Refresh
    ws["C1"] = "Last Refresh"
    ws["C1"].font = Font(name="Calibri", bold=True)
    ws["D1"] = "=NOW()"
    c.apply_calc(ws["D1"])
    ws["D1"].number_format = "yyyy-mm-dd hh:mm"

    # Header row 2
    for idx, h in enumerate(HEADERS):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}2"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    # Data rows 3..17 (15 peers default; design allows up to A2:O101)
    for r in range(3, 18):
        tk = ws[f"B{r}"]
        c.apply_input(tk)
        for col, tpl in CIQ_FORMULAS.items():
            cell = ws[f"{col}{r}"]
            cell.value = tpl.format(r=r)
            c.apply_ciq(cell)
            if col in ("E", "F", "G", "H"):
                cell.number_format = c.NUM_FMT_ACCOUNTING
            elif col == "I":
                cell.number_format = c.NUM_FMT_PERCENT
            elif col in ("J", "K", "L", "M", "N"):
                cell.number_format = c.NUM_FMT_MULTIPLE
            elif col == "O":
                cell.number_format = c.NUM_FMT_DATE

    # Currency warning P-col (P3:P17)
    for r in range(3, 18):
        cell = ws[f"P{r}"]
        cell.value = f'=IF(AND(D{r}<>"",D{r}<>"KRW"),"⚠ FX변환필요","")'
        c.apply_calc(cell)

    return ws
