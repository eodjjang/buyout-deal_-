"""9c_Manual_Supplement — Korean non-listed deals supplement."""
from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from lbo_template import conventions as c
from lbo_template.layout import SHEET_9C

# Same 15 cols as 9b + Source (col P) + Reliability (col Q) + Include (col R) + Memo (S)
HEADERS_BASE = [
    "Transaction ID (MAN-xxx)",
    "Announced Date",
    "Closed Date",
    "Target Company Name",
    "Target Country",
    "Target Primary Industry",
    "Buyer Name",
    "Buyer Type",
    "Transaction Currency",
    "Implied Enterprise Value",
    "Target LTM Revenue",
    "Target LTM EBITDA",
    "Implied EV / LTM Revenue",
    "Implied EV / LTM EBITDA",
    "Deal Status",
]
EXTRA = ["Source", "Reliability", "Include ✓", "Memo"]

SOURCE_MAP = [
    ("내부DB", "High", True),
    ("Kisvalue", "High", True),
    ("IR자료", "High", True),
    ("한경Compass", "Medium", True),
    ("투자조선", "Medium", True),
    ("루머성", "Low", False),
]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_9C)
    for col_idx in range(len(HEADERS_BASE) + len(EXTRA)):
        col = chr(ord("A") + col_idx)
        ws.column_dimensions[col].width = 16

    all_headers = HEADERS_BASE + EXTRA
    for idx, h in enumerate(all_headers):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}2"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    ws["A50"] = "Source"
    ws["A50"].font = Font(name="Calibri", bold=True)
    ws["B50"] = "Default Reliability"
    ws["B50"].font = Font(name="Calibri", bold=True)
    ws["C50"] = "Default Include"
    ws["C50"].font = Font(name="Calibri", bold=True)
    for i, (src, rel, inc) in enumerate(SOURCE_MAP):
        ws.cell(row=51 + i, column=1, value=src)
        ws.cell(row=51 + i, column=2, value=rel)
        ws.cell(row=51 + i, column=3, value=inc)

    src_opts = ",".join([s for s, _, _ in SOURCE_MAP])
    dv_src = DataValidation(type="list", formula1=f'"{src_opts}"', allow_blank=True)
    dv_rel = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    ws.add_data_validation(dv_src)
    ws.add_data_validation(dv_rel)

    for r in range(3, 50):
        for col_idx in range(len(HEADERS_BASE)):
            col = chr(ord("A") + col_idx)
            c.apply_input(ws[f"{col}{r}"])
        src_cell = f"P{r}"
        dv_src.add(src_cell)
        c.apply_input(ws[src_cell])
        rel_cell = ws[f"Q{r}"]
        # INDEX/MATCH — Excel 2016+ 호환 (XLOOKUP은 구버전에서 #NAME?)
        rel_cell.value = (
            f'=IFERROR(INDEX($B$51:$B$56,MATCH(P{r},$A$51:$A$56,0)),"")'
        )
        c.apply_calc(rel_cell)
        dv_rel.add(f"Q{r}")
        inc_cell = ws[f"R{r}"]
        inc_cell.value = (
            f'=IFERROR(INDEX($C$51:$C$56,MATCH(P{r},$A$51:$A$56,0)),FALSE)'
        )
        c.apply_calc(inc_cell)
        c.apply_input(ws[f"S{r}"])

    return ws
