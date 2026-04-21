"""0_README sheet — versioning, conventions, Preconditions, CapIQ setup guide."""
from __future__ import annotations

from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment

from lbo_template.layout import SHEET_README
from lbo_template import conventions as c


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_README)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 80

    ws["A1"] = "LBO Stress Template v0.5 — 대주단 관점 범용"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:F1")

    sections: list[tuple[str, list[str]]] = [
        ("1. Version History", [
            "v0.5 (2026-04-20) — Phase 0 Preconditions 1~3 해결, Alt-A 확정, HMM 양식 반영, Valuation 추상화",
            "v0.4 (2026-04-20) — Spec Self-Review 반영, P·Q 분기 v1.1 backlog로 연기",
            "v0.3 (2026-04-20) — CapIQ Plug-in 수식 primary, P·Q 분기 추가(추후 연기)",
            "v0.2 (2026-04-20) — CapIQ Export-once Cascade-everywhere 구조 도입",
            "v0.1 (2026-04-20) — 초안",
        ]),
        ("2. 색상·폰트·단위 컨벤션", [
            "섹션 헤더 fill: #1F4E79 / 컬럼 헤더 fill: #D9E1F2 / 입력 fill: #F2F2F2 / Key output fill: #BDD7EE",
            "폰트: 입력 #0000FF / 계산 #000000 / 동일탭 링크 #800080 / 타탭 링크 #008000 / CIQ Plug-in 수식 #008B8B",
            "단위: KRW 백만원 단일. USD 딜은 입력 전 환산. 단위 혼용 금지.",
            "Sign convention: 지출·차감 모두 양수 표기 + 수식에서 차감.",
            "시점 축: FY-2 Actual / FY-1 Actual / LTM / FY1 / FY2 / FY3 / FY4 / FY5 (8개 컬럼).",
        ]),
        ("3. Phase 0 Preconditions 체크리스트", [
            "☑ Precondition 1 — Lender-Adjusted EBITDA 정의: 본부 기준 모든 add-back 불인정. Reported EBITDA 단일.",
            "☑ Precondition 2 — Word 심사보고서 표준 양식: HMM 보고서(2023-23차) 양식 자동 추출 반영.",
            "☑ Precondition 3 — CapIQ IT/DLP 제약: Plug-in 유효 / DLP 비차단 / 호출 한도 미상(empirical 확인).",
        ]),
        ("4. CapIQ Saved Screen 사전 세팅", [
            "Saved Screen A (Trading Peers): 15개 컬럼 순서 — Company Name / CIQ ID / Country / Currency / Market Cap / EV / LTM Revenue / LTM EBITDA / LTM EBITDA Margin % / EV/LTM EBITDA / EV/FY-1 / EV/FY-2 / EV/NTM / Net Debt/LTM EBITDA / LTM Period End Date.",
            "Saved Screen B (Transaction Comps): 15개 컬럼 — Transaction ID / Announced / Closed / Target / Target Country / Primary Industry / Buyer / Buyer Type / Currency / Implied EV / LTM Rev / LTM EBITDA / EV/Rev / EV/EBITDA / Deal Status.",
            "Ticker List는 비워두고 딜마다 갈아끼움. 나머지 14개 컬럼 순서는 고정.",
            "Saved Screen URL을 본 시트 하단 '북마크' 섹션에 수기 기록.",
        ]),
        ("5. Alt-A: Paste Fallback 시 재배포 절차", [
            "설계 §9 Alt-A(동일 셀 택1) 구조상, Plug-in 비가용 시 Paste Special Values는 1회 실행하는 순간 9a/9b의 =CIQ() 수식을 영구 소실시킴.",
            "복구 절차: (1) 마스터 템플릿 파일을 사내 팀 드라이브에서 재다운로드. (2) 현재 작업 중인 입력값(1_Input, 2_Stress, 9c_Manual)만 복사해 옮김. (3) 9a/9b는 Ticker 리스트를 재입력하고 Data → Refresh All 1회.",
            "Mode 셀(9a!B1, 9b!B1)이 'Paste Fallback — 재배포 필요'로 전환되면 즉시 이 절차 실행 권장.",
        ]),
        ("6. 시트 맵", [
            "1_Input_BaseCase — 매수자 모델 4대 드라이버 + 인수조건 입력 (단일 진입점)",
            "2_Stress_Panel — Case_Switch + 6개 스트레스 파라미터",
            "3_Operating_Overlay — Stressed Revenue/EBITDA/Capex/NWC/UFCF 계산",
            "4_Debt_Schedule — Opco Senior / Opco 2nd Lien / Holdco Sub 3-트랜치",
            "5_CF_Waterfall — Opco UFCF → 이자·원금 → Dividend → Holdco",
            "6_DCF_Valuation — FCFF 5Y + Gordon TV (영구성장 1.0% 고정, 할인기간 5.0)",
            "7_Returns_LTV — 평가방식 1/2/3 추상화 + 9-열 LTV 표",
            "8_Dashboard — Word 복붙용 요약 (DASH_* Named Range cluster + 차주 자금수지표)",
            "9a_CIQ_Trading_Raw — CapIQ Trading Peer 수식 zone (Plug-in primary, Paste fallback)",
            "9b_CIQ_Transaction_Raw — CapIQ Transaction Comps 수식 zone (max 500행)",
            "9c_Manual_Supplement — Kisvalue/한경 Compass 수기 보완",
            "9_Peer_Summary — 3 소스 통합 + Include ✓ 최종 확정",
        ]),
    ]

    row = 3
    for title, lines in sections:
        ws.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
        row += 1
        for line in lines:
            ws.cell(row=row, column=2, value=line).alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        row += 1

    return ws
