"""2_Stress_Panel — Case_Switch and 7-parameter stress table with Active_* names."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment

from lbo_template.layout import SHEET_STRESS, CASE_SWITCH_CELL
from lbo_template import conventions as c

_SECTION_TITLE_FONT = Font(name="Calibri", bold=True, size=11)

PARAM_ROWS = [
    ("Revenue Growth Δ", 0.0, 0.02, -0.05, c.NUM_FMT_PERCENT, "Active_Revenue_Growth_Delta", "%"),
    ("EBITDA Margin Δ", 0.0, 0.005, -0.03, c.NUM_FMT_PERCENT, "Active_EBITDA_Margin_Delta", "%"),
    ("Capex % of Revenue Δ", 0.0, -0.005, 0.015, c.NUM_FMT_PERCENT, "Active_Capex_Pct_Delta", "%"),
    ("ΔNWC % of Revenue Δ", 0.0, 0.0, 0.01, c.NUM_FMT_PERCENT, "Active_NWC_Pct_Delta", "%"),
    ("WACC Uplift", 0, -50, 100, c.NUM_FMT_BPS, "Active_WACC_Uplift", "bp"),
    ("Exit Multiple Δ", 0.0, 1.0, -1.0, c.NUM_FMT_MULTIPLE, "Active_Exit_Multiple_Delta", "x"),
    ("Permanent Growth (고정)", 0.01, 0.01, 0.01, c.NUM_FMT_PERCENT, "Perm_Growth", "%"),
]

INDUSTRY_PRESETS = ["(수동)", "소매", "제조", "SaaS", "헬스케어", "해운·시황주"]


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_STRESS)
    ws.column_dimensions["A"].width = 30
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 14

    ws["A1"] = "2. Stress Panel — 시나리오 스위치 & 파라미터"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:F1")

    # --- Case_Switch (B3) -------------------------------------------------
    ws["A3"] = "Case_Switch"
    ws["A3"].font = _SECTION_TITLE_FONT
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")

    b3 = ws["B3"]
    b3.value = "Base"
    c.apply_input(b3)

    dv_case = DataValidation(type="list", formula1='"Base,Upside,Downside"', allow_blank=False)
    ws.add_data_validation(dv_case)
    dv_case.add(CASE_SWITCH_CELL)

    wb.defined_names["Case_Switch"] = DefinedName(
        "Case_Switch", attr_text=f"'{SHEET_STRESS}'!$B$3"
    )

    # --- Industry preset (reference dropdown) -----------------------------
    ws["A4"] = "업종 프리셋 (참고)"
    ws["A4"].font = _SECTION_TITLE_FONT
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    b4 = ws["B4"]
    b4.value = INDUSTRY_PRESETS[0]
    c.apply_input(b4)

    preset_list = ",".join(INDUSTRY_PRESETS)
    dv_preset = DataValidation(type="list", formula1=f'"{preset_list}"', allow_blank=True)
    ws.add_data_validation(dv_preset)
    dv_preset.add("B4")

    # --- Parameter table header (row 7) -----------------------------------
    header_cells = [
        ("A7", "파라미터"),
        ("B7", "Base"),
        ("C7", "Upside"),
        ("D7", "Downside"),
        ("E7", "단위"),
        ("F7", "Active"),
    ]
    for addr, text in header_cells:
        cell = ws[addr]
        cell.value = text
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    first_data_row = 8
    for idx, row_spec in enumerate(PARAM_ROWS):
        label, base_v, up_v, down_v, num_fmt, active_name, unit_label = row_spec
        r = first_data_row + idx

        ws.cell(row=r, column=1, value=label)
        label_cell = ws.cell(row=r, column=1)
        label_cell.font = c.calc_font()
        label_cell.border = c.thin_border()

        is_perm = "Permanent Growth" in label
        for col_letter, val in (("B", base_v), ("C", up_v), ("D", down_v)):
            cell = ws[f"{col_letter}{r}"]
            cell.value = val
            if is_perm:
                c.apply_calc(cell)
            else:
                c.apply_input(cell)
            cell.number_format = num_fmt

        unit_cell = ws.cell(row=r, column=5, value=unit_label)
        c.apply_calc(unit_cell)

        f_cell = ws[f"F{r}"]
        f_cell.value = (
            f'=SWITCH(Case_Switch,"Base",B{r},"Upside",C{r},"Downside",D{r})'
        )
        c.apply_key_output(f_cell)
        f_cell.number_format = num_fmt

        wb.defined_names[active_name] = DefinedName(
            active_name, attr_text=f"'{SHEET_STRESS}'!$F${r}"
        )

    return ws
