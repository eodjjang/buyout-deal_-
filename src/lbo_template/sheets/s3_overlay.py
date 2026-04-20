"""3_Operating_Overlay — stressed operating metrics cascade (design §3).

Row layout (overlay sheet): Base Rev r5 … UFCF r19. Template row numbers in the
original plan snippet were inconsistent; formulas below reference this fixed grid.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from lbo_template.layout import SHEET_OVERLAY, SHEET_INPUT
from lbo_template import conventions as c

ROWS = [
    ("Base Revenue", "='{inp}'!{c}24"),
    ("Base YoY Growth", "=IFERROR({c}5/{prev}5-1,\"\")"),
    (
        "Stressed YoY Growth",
        "=IF({is_forecast},{c}6+Active_Revenue_Growth_Delta,{c}6)",
    ),
    ("Stressed Revenue", "=IF({is_forecast},{prev}8*(1+{c}7),{c}5)"),
    (
        "Base EBITDA Margin",
        "=IFERROR('{inp}'!{c}27/'{inp}'!{c}24,\"\")",
    ),
    (
        "Stressed EBITDA Margin",
        "=IF({is_forecast},{c}9+Active_EBITDA_Margin_Delta,{c}9)",
    ),
    ("Stressed EBITDA", "={c}8*{c}10"),
    (
        "Base Capex % of Revenue",
        "=IFERROR('{inp}'!{c}29/'{inp}'!{c}24,\"\")",
    ),
    (
        "Stressed Capex",
        "={c}8*({c}12+IF({is_forecast},Active_Capex_Pct_Delta,0))",
    ),
    (
        "Base ΔNWC % of Revenue",
        "=IFERROR('{inp}'!{c}30/'{inp}'!{c}24,\"\")",
    ),
    (
        "Stressed ΔNWC",
        "={c}8*({c}14+IF({is_forecast},Active_NWC_Pct_Delta,0))",
    ),
    ("D&A (Base pass-through)", "='{inp}'!{c}28"),
    ("EBIT (Stressed)", "={c}11-{c}16"),
    ("Cash Taxes", "=MAX(0,{c}17)*'{inp}'!{c}31"),
    ("UFCF (Stressed)", "={c}11-{c}18-{c}13-{c}15"),
]

FIRST_DATA_ROW = 5
# Row index of the last overlay line (UFCF); used by 5_CF_Waterfall and downstream sheets.
UFCF_ROW = FIRST_DATA_ROW + len(ROWS) - 1
# "Stressed EBITDA" is index 6 in ROWS → row 11 (not Capex row 13).
STRESSED_EBITDA_ROW = FIRST_DATA_ROW + 6

FORECAST_COLS = {"E", "F", "G", "H", "I"}


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_OVERLAY)
    ws.column_dimensions["A"].width = 32
    for col in c.FY_AXIS_COLUMNS:
        ws.column_dimensions[col].width = 14

    ws["A1"] = "3. Operating Overlay (Stressed)"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:I1")

    for col, label in zip(c.FY_AXIS_COLUMNS, c.FY_AXIS_LABELS):
        cell = ws[f"{col}4"]
        cell.value = label
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    for idx, (label, template) in enumerate(ROWS):
        r = 5 + idx
        ws.cell(row=r, column=1, value=label)
        for col_idx, col in enumerate(c.FY_AXIS_COLUMNS):
            cell = ws[f"{col}{r}"]
            if col_idx == 0:
                prev = col
            else:
                prev = c.FY_AXIS_COLUMNS[col_idx - 1]
            is_forecast = "TRUE" if col in FORECAST_COLS else "FALSE"
            formula = template.format(
                c=col,
                prev=prev,
                inp=SHEET_INPUT,
                is_forecast=is_forecast,
            )
            cell.value = formula
            if label.startswith("Base "):
                c.apply_calc(cell)
                cell.font = c.crosstab_link_font()
            elif label.startswith("Stressed ") or label in (
                "UFCF (Stressed)",
                "EBIT (Stressed)",
            ):
                c.apply_key_output(cell)
            else:
                c.apply_calc(cell)
            if (
                "Margin" in label
                or "Growth" in label
                or "% of Revenue" in label
            ):
                cell.number_format = c.NUM_FMT_PERCENT
            else:
                cell.number_format = c.NUM_FMT_ACCOUNTING

    return ws
