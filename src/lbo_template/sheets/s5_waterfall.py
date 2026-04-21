"""5_CF_Waterfall — Opco UFCF → dividend → Holdco (design §5)."""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

from lbo_template.layout import SHEET_WATERFALL, SHEET_OVERLAY, SHEET_DEBT
from lbo_template import conventions as c
from lbo_template.sheets.s3_overlay import STRESSED_EBITDA_ROW, UFCF_ROW

_RED_FILL = PatternFill("solid", fgColor="FFC7CE")

# Main waterfall block: row 5 = first line (aligned with plan Task 9).
ROWS: list[tuple[str, str | None]] = [
    ("Opco UFCF", None),  # set per column: overlay UFCF
    ("Less: Opco Interest (Senior + 2nd Lien)", None),
    ("Less: Opco Mandatory Amort", None),
    ("= Opco CFADS", "={c}5-{c}6-{c}7"),
    ("Less: Minimum Cash Retention", None),
    ("Less: Legal Reserve", None),
    ("= Distributable to Holdco", "=MAX(0,{c}8-{c}9-{c}10)"),
    ("× Payout Ratio", None),
    ("= Dividend Paid to Holdco", "={c}11*{c}12"),
    ("Opco Sweep Available", "=MAX(0,{c}8-{c}13)*Sweep_Pct"),
    ("Holdco Dividend Received", "={c}13"),
    ("Holdco Interest (if Cash-Pay)", None),
    ("Holdco Net Cash Flow", "={c}15-{c}16"),
    ("Holdco ICR (Div / Holdco Interest)", '=IFERROR({c}15/{c}16,"")'),
]

SWEEP_ROW = 5 + 9  # index 9 → Opco Sweep Available
KPI_START = 21


def _opco_ufcf_formula(col: str) -> str:
    return f"='{SHEET_OVERLAY}'!{col}{UFCF_ROW}"


def _opco_interest_formula(col: str) -> str:
    return f"='{SHEET_DEBT}'!{col}8+'{SHEET_DEBT}'!{col}18"


def _opco_mand_formula(col: str) -> str:
    return f"='{SHEET_DEBT}'!{col}9+'{SHEET_DEBT}'!{col}19"


def _holdco_interest_formula(col: str) -> str:
    return f"=IF(Holdco_PIK_Mode=\"Cash\",'{SHEET_DEBT}'!{col}28,0)"


def _kpi_dscr(col: str) -> str:
    return f'=IFERROR({col}8/({col}6+{col}7),\"\")'


def _kpi_opco_icr(col: str) -> str:
    return f"=IFERROR('{SHEET_OVERLAY}'!{col}{STRESSED_EBITDA_ROW}/{col}6,\"\")"


def _kpi_holdco_icr_dup(col: str) -> str:
    return f"={col}18"


def _kpi_net_leverage(col: str) -> str:
    # Debt ending balances: Senior r11, 2nd r21, Holdco r31 (see s4_debt TRANCHE_ROWS).
    return (
        f"=IFERROR(('{SHEET_DEBT}'!{col}11+'{SHEET_DEBT}'!{col}21+'{SHEET_DEBT}'!{col}31)/"
        f"'{SHEET_OVERLAY}'!{col}{STRESSED_EBITDA_ROW},\"\")"
    )


KPI_ROWS: list[tuple[str, str]] = [
    ("Opco DSCR (CFADS / (Int+Mand))", "dscr"),
    ("Opco ICR (EBITDA / Opco Int)", "opco_icr"),
    ("Holdco ICR", "holdco_icr"),
    ("Net Leverage ((Opco+Holdco)/EBITDA)", "net_lev"),
]

_KPI_BUILDERS = {
    "dscr": _kpi_dscr,
    "opco_icr": _kpi_opco_icr,
    "holdco_icr": _kpi_holdco_icr_dup,
    "net_lev": _kpi_net_leverage,
}


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_WATERFALL)
    ws.column_dimensions["A"].width = 44
    for col in c.FY_AXIS_COLUMNS:
        ws.column_dimensions[col].width = 14

    ws["A1"] = "5. Cash Flow Waterfall"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:I1")

    for col, label in zip(c.FY_AXIS_COLUMNS, c.FY_AXIS_LABELS):
        hdr = ws[f"{col}3"]
        hdr.value = label
        hdr.font = c.column_header_font()
        hdr.fill = c.column_header_fill()

    for idx, (label, template) in enumerate(ROWS):
        r = 5 + idx
        ws.cell(row=r, column=1, value=label)
        for col_idx, col in enumerate(c.FY_AXIS_COLUMNS):
            cell = ws[f"{col}{r}"]
            if col_idx < 3:
                continue
            if template is None:
                if label == "Opco UFCF":
                    cell.value = _opco_ufcf_formula(col)
                elif label == "Less: Opco Interest (Senior + 2nd Lien)":
                    cell.value = _opco_interest_formula(col)
                elif label == "Less: Opco Mandatory Amort":
                    cell.value = _opco_mand_formula(col)
                elif label == "Holdco Interest (if Cash-Pay)":
                    cell.value = _holdco_interest_formula(col)
                elif label == "Less: Minimum Cash Retention":
                    cell.value = 0
                    cell.number_format = c.NUM_FMT_ACCOUNTING
                    c.apply_input(cell)
                elif label == "Less: Legal Reserve":
                    cell.value = 0
                    cell.number_format = c.NUM_FMT_ACCOUNTING
                    c.apply_input(cell)
                elif label == "× Payout Ratio":
                    cell.value = 1.0
                    cell.number_format = c.NUM_FMT_PERCENT
                    c.apply_input(cell)
                else:
                    cell.value = 0
                    c.apply_input(cell)
                continue

            cell.value = template.format(c=col)
            c.apply_calc(cell)
            cell.number_format = c.NUM_FMT_ACCOUNTING
            if label.startswith("=") or "Dividend" in label or "CFADS" in label:
                c.apply_key_output(cell)
            elif "ICR" in label and "Holdco ICR" in label:
                cell.number_format = c.NUM_FMT_MULTIPLE
                c.apply_key_output(cell)

    for col in ["E", "F", "G", "H", "I"]:
        name = f"Opco_Sweep_Avail_{col}"
        c.define_name(wb, name, f"'{SHEET_WATERFALL}'!${col}${SWEEP_ROW}")

    for idx, (label, key) in enumerate(KPI_ROWS):
        r = KPI_START + idx
        ws.cell(row=r, column=1, value=label)
        builder = _KPI_BUILDERS[key]
        for col_idx, col in enumerate(c.FY_AXIS_COLUMNS):
            if col_idx < 3:
                continue
            cell = ws[f"{col}{r}"]
            cell.value = builder(col)
            c.apply_key_output(cell)
            if "Leverage" in label or "ICR" in label or "DSCR" in label:
                cell.number_format = c.NUM_FMT_MULTIPLE
            else:
                cell.number_format = c.NUM_FMT_ACCOUNTING

    c.define_name(
        wb,
        "Opco_DSCR_Row",
        f"'{SHEET_WATERFALL}'!$E${KPI_START}:$I${KPI_START}",
    )
    c.define_name(
        wb,
        "Opco_ICR_Row",
        f"'{SHEET_WATERFALL}'!$E${KPI_START + 1}:$I${KPI_START + 1}",
    )
    c.define_name(
        wb,
        "Holdco_ICR_Row",
        f"'{SHEET_WATERFALL}'!$E${KPI_START + 2}:$I${KPI_START + 2}",
    )
    c.define_name(
        wb,
        "Net_Leverage_Row",
        f"'{SHEET_WATERFALL}'!$E${KPI_START + 3}:$I${KPI_START + 3}",
    )
    c.define_name(wb, "Dividend_Row", f"'{SHEET_WATERFALL}'!$E$13:$I$13")

    ws.conditional_formatting.add(
        f"E18:I18",
        CellIsRule(operator="lessThan", formula=["1"], fill=_RED_FILL),
    )

    return ws
