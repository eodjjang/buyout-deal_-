"""7_Returns_LTV — abstracted Valuation Method 1/2/3 per design v0.5."""
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
from lbo_template.layout import SHEET_RETURNS
from lbo_template import conventions as c

METHOD_TYPES = [
    "DCF_Stressed",
    "Trading_EV_EBITDA",
    "Trading_PBR",
    "Trading_PER",
    "Transaction_EV_EBITDA",
    "Transaction_PBR",
    "MarketCap_Avg",
    "Manual_Absolute",
]


def _if_chain_multiple(cr: str) -> str:
    """Excel 2016 이전 호환 — SWITCH 대신 IF 중첩."""
    return (
        f'=IF({cr}="DCF_Stressed",IFERROR(DCF_EV/D27,"n.a"),'
        f'IF({cr}="Trading_EV_EBITDA",Applied_Trading_Multiple,'
        f'IF({cr}="Trading_PBR",Applied_Trading_PBR,'
        f'IF({cr}="Trading_PER","(수기)",'
        f'IF({cr}="Transaction_EV_EBITDA",Applied_Transaction_Multiple,'
        f'IF({cr}="Transaction_PBR","(수기)",'
        f'IF({cr}="MarketCap_Avg","n.a",'
        f'IF({cr}="Manual_Absolute","n.a",""))))))))'
    )


def _if_chain_base_metric(cr: str) -> str:
    return (
        f'=IF({cr}="DCF_Stressed","EBITDA",'
        f'IF({cr}="Trading_EV_EBITDA","EBITDA",'
        f'IF({cr}="Trading_PBR","Book Value",'
        f'IF({cr}="Trading_PER","Net Income",'
        f'IF({cr}="Transaction_EV_EBITDA","EBITDA",'
        f'IF({cr}="Transaction_PBR","Book Value",'
        f'IF({cr}="MarketCap_Avg","절대값",'
        f'IF({cr}="Manual_Absolute","절대값",""))))))))'
    )


def _if_chain_ltv_ev(method_r: int) -> str:
    cr = f"C{method_r}"
    return (
        f'=IFERROR(IF({cr}="DCF_Stressed",DCF_Equity_Value,'
        f'IF({cr}="MarketCap_Avg",F{method_r},'
        f'IF({cr}="Manual_Absolute",F{method_r},'
        f'D{method_r}*LTM_EBITDA))),"")'
    )


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_RETURNS)
    ws.column_dimensions["A"].width = 14
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 15

    ws["A1"] = "7. Returns & LTV — Valuation Method 1/2/3 (v0.5 추상화)"
    c.apply_section_header(ws["A1"])
    ws.merge_cells("A1:K1")

    # --- Method Declaration Block (header row 3, data rows 4..6) ---
    decl_headers = ["방식", "Label", "Method Type", "Multiple", "Multiple 기반", "Source 메모"]
    for idx, h in enumerate(decl_headers):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}3"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    DEFAULTS = [
        ("평가방식 1", "DCF (Stressed)", "DCF_Stressed"),
        ("평가방식 2", "Trading EV/EBITDA", "Trading_EV_EBITDA"),
        ("평가방식 3", "Transaction EV/EBITDA (Trimmed)", "Transaction_EV_EBITDA"),
    ]

    dv_mt = DataValidation(type="list", formula1=f'"{",".join(METHOD_TYPES)}"', allow_blank=False)
    ws.add_data_validation(dv_mt)

    for idx, (method_tag, default_label, default_type) in enumerate(DEFAULTS):
        r = 4 + idx
        ws[f"A{r}"] = method_tag
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"B{r}"] = default_label
        c.apply_input(ws[f"B{r}"])
        ws[f"C{r}"] = default_type
        c.apply_input(ws[f"C{r}"])
        dv_mt.add(f"C{r}")

        # Multiple (col D) — Method Type 분기 (SWITCH 대신 IF: 구버전 Excel 호환)
        mul = ws[f"D{r}"]
        mul.value = _if_chain_multiple(f"C{r}")
        c.apply_calc(mul)
        mul.number_format = c.NUM_FMT_MULTIPLE
        c.define_name(wb, f"DASH_Valuation_Method{idx+1}_Label", f"'{SHEET_RETURNS}'!$B${r}")
        c.define_name(wb, f"DASH_Valuation_Method{idx+1}_Multiple", f"'{SHEET_RETURNS}'!$D${r}")

        base = ws[f"E{r}"]
        base.value = _if_chain_base_metric(f"C{r}")
        c.apply_calc(base)

        # Source memo F
        c.apply_input(ws[f"F{r}"])

    # --- 9-Column LTV Table (title row 9, header row 10, data rows 11..13) ---
    ws["A9"] = "9-열 LTV 산출"
    ws["A9"].font = Font(bold=True, size=11)

    ltv_headers = [
        "방식", "(a) 기준지표", "(b) Multiple", "(c) 지분가치 100%",
        "(d) 지분율", "(e) 담보지분가치", "(f) Opco 차입금",
        "(g) Opco LTV", "(h) Holdco 차입금", "(i) 누적 LTV",
    ]
    for idx, h in enumerate(ltv_headers):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}10"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    # LTM EBITDA anchor for implied multiple back-calc
    ws["A27"] = "LTM EBITDA (for implied multiple back-calc)"
    ws["D27"] = "=LTM_EBITDA"
    c.apply_calc(ws["D27"])

    ws["A28"] = "Target Ownership (지분율)"
    ws["B28"] = 1.0
    c.apply_input(ws["B28"])
    ws["B28"].number_format = c.NUM_FMT_PERCENT
    c.define_name(wb, "Target_Ownership", f"'{SHEET_RETURNS}'!$B$28")

    for idx in range(3):
        r = 11 + idx
        method_r = 4 + idx
        ws[f"A{r}"] = f"=B{method_r}"
        ws[f"A{r}"].font = c.sametab_link_font()
        ws[f"B{r}"] = f"=E{method_r}"
        ws[f"B{r}"].font = c.sametab_link_font()
        ws[f"C{r}"] = f"=D{method_r}"
        ws[f"C{r}"].font = c.sametab_link_font()
        # (c) 지분가치 100% — dispatched by Method Type
        ws[f"D{r}"] = _if_chain_ltv_ev(method_r)
        c.apply_calc(ws[f"D{r}"])
        ws[f"D{r}"].number_format = c.NUM_FMT_ACCOUNTING
        ws[f"E{r}"] = "=Target_Ownership"
        ws[f"E{r}"].number_format = c.NUM_FMT_PERCENT
        ws[f"F{r}"] = f"=D{r}*E{r}"
        c.apply_key_output(ws[f"F{r}"])
        ws[f"F{r}"].number_format = c.NUM_FMT_ACCOUNTING
        ws[f"G{r}"] = "=Opco_Senior_Principal+Opco_2L_Principal"
        ws[f"G{r}"].number_format = c.NUM_FMT_ACCOUNTING
        ws[f"H{r}"] = f'=IFERROR(G{r}/F{r},"")'
        c.apply_key_output(ws[f"H{r}"])
        ws[f"H{r}"].number_format = c.NUM_FMT_PERCENT
        ws[f"I{r}"] = "=Holdco_Sub_Principal"
        ws[f"I{r}"].number_format = c.NUM_FMT_ACCOUNTING
        ws[f"J{r}"] = f'=IFERROR((G{r}+I{r})/F{r},"")'
        c.apply_key_output(ws[f"J{r}"])
        ws[f"J{r}"].number_format = c.NUM_FMT_PERCENT

        c.define_name(wb, f"DASH_Valuation_Method{idx+1}_EV", f"'{SHEET_RETURNS}'!$D${r}")
        c.define_name(wb, f"DASH_LTV_Method{idx+1}_Opco", f"'{SHEET_RETURNS}'!$H${r}")
        c.define_name(wb, f"DASH_LTV_Method{idx+1}_Cumulative", f"'{SHEET_RETURNS}'!$J${r}")

    return ws
