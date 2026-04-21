"""9b_CIQ_Transaction_Raw — Transaction Comps zone (500 rows)."""
from __future__ import annotations

from openpyxl.styles import Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from lbo_template import conventions as c
from lbo_template.layout import SHEET_9B, S9B_DATA_END_ROW

HEADERS = [
    "Transaction ID",
    "Announced Date",
    "Closed Date",
    "Target Company Name",
    "Target Country",
    "Target Primary Industry",
    "Buyer Name",
    "Buyer Type",
    "Transaction Currency",
    "Implied Enterprise Value",
    "Target LTM Revenue",
    "Target LTM EBITDA",
    "Implied EV / LTM Revenue",
    "Implied EV / LTM EBITDA",
    "Deal Status",
]

# A열 = CIQ 거래 ID(IQTR…). S&P 매뉴얼 예: =CIQ("IQTR…","IQ_TR_TARGETNAME").
# 그 외 IQ_TR_* 는 Formula Builder → Transactions 폴더에서 항목명이 다를 수 있음(플러그인 버전).
_TX_DATA_ITEMS: dict[str, str] = {
    "B": "IQ_TR_ANNOUNCED_DATE",
    "C": "IQ_TR_CLOSED_DATE",
    "D": "IQ_TR_TARGETNAME",
    "E": "IQ_TR_TARGET_COUNTRY",
    "F": "IQ_TR_TARGET_INDUSTRY",
    "G": "IQ_TR_BUYER",
    "H": "IQ_TR_BUYER_TYPE",
    "I": "IQ_TR_CURRENCY",
    "J": "IQ_TR_IMPLIED_EV",
    "K": "IQ_TR_LTM_REVENUE",
    "L": "IQ_TR_LTM_EBITDA",
    "M": "IQ_TR_IMPLIED_EV_TO_LTM_REV",
    "N": "IQ_TR_IMPLIED_EV_TO_LTM_EBITDA",
    "O": "IQ_TR_STATUS",
}


def _tx_cell(r: int, col_letter: str) -> str:
    """Return Excel formula for transaction row r, column letter (B..O)."""
    item = _TX_DATA_ITEMS[col_letter]
    return f'=IF($A{r}="","",IFERROR(CIQ($A{r},"{item}"),""))'


def build(wb: Workbook) -> Worksheet:
    ws = wb.create_sheet(SHEET_9B)
    for col_idx in range(15):
        col = chr(ord("A") + col_idx)
        ws.column_dimensions[col].width = 16

    ws["A1"] = "Mode"
    ws["A1"].font = Font(name="Calibri", bold=True)
    ws["B1"] = '=IF(AND(ISFORMULA(D3),$A3<>""),"Plug-in","⚠ Paste Fallback — 마스터 재배포 필요")'
    c.apply_calc(ws["B1"])

    ws["C1"] = (
        '=IF(COUNTA(A:A)-1>500,"⚠ Export 500행 초과 — Paste 잘림 위험. 필터 좁히거나 범위 확장 필요","OK")'
    )
    c.apply_calc(ws["C1"])

    for idx, h in enumerate(HEADERS):
        col = chr(ord("A") + idx)
        cell = ws[f"{col}2"]
        cell.value = h
        cell.font = c.column_header_font()
        cell.fill = c.column_header_fill()

    for r in range(3, S9B_DATA_END_ROW + 1):
        a_cell = ws[f"A{r}"]
        c.apply_input(a_cell)
        for col in "BCDEFGHIJKLMNO":
            cell = ws[f"{col}{r}"]
            cell.value = _tx_cell(r, col)
            c.apply_ciq(cell)
        ws[f"B{r}"].number_format = c.NUM_FMT_DATE
        ws[f"C{r}"].number_format = c.NUM_FMT_DATE
        for col in "JKL":
            ws[f"{col}{r}"].number_format = c.NUM_FMT_ACCOUNTING
        for col in "MN":
            ws[f"{col}{r}"].number_format = c.NUM_FMT_MULTIPLE

    return ws
