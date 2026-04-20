"""Style, color, font, and number-format conventions (design doc §0)."""
from __future__ import annotations
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLOR_SECTION_HEADER_FILL = "1F4E79"
COLOR_COLUMN_HEADER_FILL = "D9E1F2"
COLOR_INPUT_FILL = "F2F2F2"
COLOR_KEY_OUTPUT_FILL = "BDD7EE"

COLOR_INPUT_FONT = "0000FF"
COLOR_CALC_FONT = "000000"
COLOR_SAMETAB_LINK_FONT = "800080"
COLOR_CROSSTAB_LINK_FONT = "008000"
COLOR_CIQ_FORMULA_FONT = "008B8B"
COLOR_WHITE_FONT = "FFFFFF"

NUM_FMT_ACCOUNTING = '$#,##0;($#,##0);"-"'
NUM_FMT_PERCENT = "0.0%"
NUM_FMT_PERCENT_PP = '0.0"%p"'
NUM_FMT_MULTIPLE = '0.0"x"'
NUM_FMT_BPS = '0" bp"'
NUM_FMT_DATE = "yyyy-mm-dd"

FY_AXIS_COLUMNS = ["B", "C", "D", "E", "F", "G", "H", "I"]
FY_AXIS_LABELS = ["FY-2", "FY-1", "LTM", "FY1", "FY2", "FY3", "FY4", "FY5"]
FY_AXIS_INDEX = dict(zip(FY_AXIS_LABELS, FY_AXIS_COLUMNS))


def section_header_font() -> Font:
    return Font(name="Calibri", size=11, bold=True, color=COLOR_WHITE_FONT)


def column_header_font() -> Font:
    return Font(name="Calibri", size=10, bold=True, color=COLOR_CALC_FONT)


def input_font() -> Font:
    return Font(name="Calibri", size=10, color=COLOR_INPUT_FONT)


def calc_font() -> Font:
    return Font(name="Calibri", size=10, color=COLOR_CALC_FONT)


def sametab_link_font() -> Font:
    return Font(name="Calibri", size=10, color=COLOR_SAMETAB_LINK_FONT)


def crosstab_link_font() -> Font:
    return Font(name="Calibri", size=10, color=COLOR_CROSSTAB_LINK_FONT)


def ciq_formula_font() -> Font:
    return Font(name="Calibri", size=10, color=COLOR_CIQ_FORMULA_FONT)


def section_header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_SECTION_HEADER_FILL)


def column_header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_COLUMN_HEADER_FILL)


def input_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_INPUT_FILL)


def key_output_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COLOR_KEY_OUTPUT_FILL)


def thin_border() -> Border:
    side = Side(style="thin", color="BFBFBF")
    return Border(left=side, right=side, top=side, bottom=side)


def apply_section_header(cell) -> None:
    cell.font = section_header_font()
    cell.fill = section_header_fill()
    cell.alignment = Alignment(horizontal="left", vertical="center")


def apply_input(cell) -> None:
    cell.font = input_font()
    cell.fill = input_fill()
    cell.border = thin_border()


def apply_calc(cell) -> None:
    cell.font = calc_font()
    cell.border = thin_border()


def apply_key_output(cell) -> None:
    cell.font = calc_font()
    cell.fill = key_output_fill()
    cell.border = thin_border()


def apply_ciq(cell) -> None:
    cell.font = ciq_formula_font()
    cell.border = thin_border()
